from flask import Flask, request, redirect, session, send_file
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)
app.secret_key = "estoque123"

def conectar_banco():
    return sqlite3.connect("database.db")

def criar_tabelas():
    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS produtos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL UNIQUE,
        quantidade INTEGER NOT NULL
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS movimentacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        produto TEXT NOT NULL,
        tipo TEXT NOT NULL,
        quantidade INTEGER NOT NULL,
        data TEXT NOT NULL
    )
    """)

    conexao.commit()
    conexao.close()

def pegar_data(data_formulario):
    if data_formulario == "":
        return datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    return data_formulario

def verificar_login():
    return session.get("logado") == True

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        senha = request.form["senha"]

        if usuario == "admin" and senha == "123":
            session["logado"] = True
            return redirect("/")
        else:
            return "Usuário ou senha incorretos. <a href='/login'>Voltar</a>"

    return """
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <title>Login</title>
        <link rel="stylesheet" href="/static/style.css">
    </head>
    <body>
        <div class="container">
            <div class="card">
                <h1>Login</h1>
                <form method="POST">
                    <input type="text" name="usuario" placeholder="Usuário" required>
                    <input type="password" name="senha" placeholder="Senha" required>
                    <button type="submit">Entrar</button>
                </form>
                <p>Usuário: admin | Senha: 123</p>
            </div>
        </div>
    </body>
    </html>
    """

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/")
def home():
    if not verificar_login():
        return redirect("/login")

    criar_tabelas()

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("SELECT * FROM produtos")
    produtos = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM produtos")
    total_produtos = cursor.fetchone()[0]

    cursor.execute("SELECT SUM(quantidade) FROM produtos")
    total_estoque = cursor.fetchone()[0] or 0

    cursor.execute("SELECT COUNT(*) FROM movimentacoes WHERE tipo = 'Entrada'")
    total_entradas = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM movimentacoes WHERE tipo = 'Saída'")
    total_saidas = cursor.fetchone()[0]

    html = f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <title>Sistema de Estoque</title>
        <link rel="stylesheet" href="/static/style.css">
    </head>
    <body>
        <div class="container">
            <h1>Sistema de Controle de Estoque</h1>

            <a href="/relatorio"><button>Ver Relatório</button></a>
            <a href="/exportar_excel"><button>Exportar Excel</button></a>
            <a href="/logout"><button class="btn-excluir">Sair</button></a>

            <div class="dashboard">
                <div class="card">
                    <h2>{total_produtos}</h2>
                    <p>Produtos cadastrados</p>
                </div>

                <div class="card">
                    <h2>{total_estoque}</h2>
                    <p>Itens em estoque</p>
                </div>

                <div class="card">
                    <h2>{total_entradas}</h2>
                    <p>Entradas registradas</p>
                </div>

                <div class="card">
                    <h2>{total_saidas}</h2>
                    <p>Saídas registradas</p>
                </div>
            </div>

            <div class="card">
                <h2>Entrada de Produtos</h2>
                <form action="/salvar" method="POST">
                    <input type="text" name="nome" placeholder="Nome do produto" required>
                    <input type="number" name="quantidade" placeholder="Quantidade" required>
                    <input type="datetime-local" name="data">
                    <button type="submit">Adicionar</button>
                </form>
            </div>

            <div class="card">
                <h2>Saída de Produtos</h2>
                <form action="/saida" method="POST">
                    <input type="text" name="nome" placeholder="Nome do produto" required>
                    <input type="number" name="quantidade" placeholder="Quantidade" required>
                    <input type="datetime-local" name="data">
                    <button type="submit">Retirar</button>
                </form>
            </div>

            <div class="card">
                <h2>Produtos em Estoque</h2>
                <ul>
    """

    for produto in produtos:
        html += f"""
        <li>
            {produto[1]} - {produto[2]} unidades
            <a href="/excluir/{produto[0]}">
                <button class="btn-excluir">Excluir</button>
            </a>
        </li>
        """

    html += """
                </ul>
            </div>
        </div>
    </body>
    </html>
    """

    conexao.close()
    return html

@app.route("/salvar", methods=["POST"])
def salvar():
    if not verificar_login():
        return redirect("/login")

    criar_tabelas()

    nome = request.form["nome"]
    quantidade = int(request.form["quantidade"])
    data_atual = pegar_data(request.form["data"])

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("SELECT * FROM produtos WHERE nome = ?", (nome,))
    produto_existente = cursor.fetchone()

    if produto_existente:
        nova_quantidade = produto_existente[2] + quantidade
        cursor.execute("UPDATE produtos SET quantidade = ? WHERE id = ?", (nova_quantidade, produto_existente[0]))
    else:
        cursor.execute("INSERT INTO produtos (nome, quantidade) VALUES (?, ?)", (nome, quantidade))

    cursor.execute("""
    INSERT INTO movimentacoes (produto, tipo, quantidade, data)
    VALUES (?, ?, ?, ?)
    """, (nome, "Entrada", quantidade, data_atual))

    conexao.commit()
    conexao.close()

    return redirect("/")

@app.route("/saida", methods=["POST"])
def saida():
    if not verificar_login():
        return redirect("/login")

    criar_tabelas()

    nome = request.form["nome"]
    quantidade = int(request.form["quantidade"])
    data_atual = pegar_data(request.form["data"])

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("SELECT * FROM produtos WHERE nome = ?", (nome,))
    produto = cursor.fetchone()

    if produto:
        nova_quantidade = produto[2] - quantidade

        if nova_quantidade < 0:
            nova_quantidade = 0

        cursor.execute("UPDATE produtos SET quantidade = ? WHERE id = ?", (nova_quantidade, produto[0]))

        cursor.execute("""
        INSERT INTO movimentacoes (produto, tipo, quantidade, data)
        VALUES (?, ?, ?, ?)
        """, (nome, "Saída", quantidade, data_atual))

    conexao.commit()
    conexao.close()

    return redirect("/")

@app.route("/relatorio")
def relatorio():
    if not verificar_login():
        return redirect("/login")

    criar_tabelas()

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("SELECT * FROM movimentacoes ORDER BY id DESC")
    movimentacoes = cursor.fetchall()

    html = """
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <title>Relatório</title>
        <link rel="stylesheet" href="/static/style.css">
    </head>
    <body>
        <div class="container">
            <h1>Relatório de Entradas e Saídas</h1>

            <a href="/"><button>Voltar</button></a>
            <a href="/exportar_excel"><button>Exportar Excel</button></a>
            <a href="/limpar_relatorio"><button class="btn-excluir">Limpar Relatório</button></a>

            <div class="card">
                <table>
                    <tr>
                        <th>Data</th>
                        <th>Produto</th>
                        <th>Tipo</th>
                        <th>Quantidade Movimentada</th>
                        <th>Estoque Atual</th>
                    </tr>
    """

    for mov in movimentacoes:
        cursor.execute("SELECT quantidade FROM produtos WHERE nome = ?", (mov[1],))
        estoque_atual = cursor.fetchone()
        estoque = estoque_atual[0] if estoque_atual else 0

        html += f"""
                    <tr>
                        <td>{mov[4]}</td>
                        <td>{mov[1]}</td>
                        <td>{mov[2]}</td>
                        <td>{mov[3]}</td>
                        <td>{estoque}</td>
                    </tr>
        """

    html += """
                </table>
            </div>
        </div>
    </body>
    </html>
    """

    conexao.close()
    return html

@app.route("/exportar_excel")
def exportar_excel():
    if not verificar_login():
        return redirect("/login")

    criar_tabelas()

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("SELECT * FROM movimentacoes ORDER BY id DESC")
    movimentacoes = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Relatório de Controle de Estoque"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F2937")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    cabecalhos = ["Data", "Produto", "Tipo", "Quantidade Movimentada", "Estoque Atual"]
    ws.append([])
    ws.append(cabecalhos)

    linha_cabecalho = 4

    for mov in movimentacoes:
        cursor.execute("SELECT quantidade FROM produtos WHERE nome = ?", (mov[1],))
        estoque_atual = cursor.fetchone()
        estoque = estoque_atual[0] if estoque_atual else 0

        ws.append([mov[4], mov[1], mov[2], mov[3], estoque])

    preenchimento_cabecalho = PatternFill("solid", fgColor="2563EB")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    borda_fina = Side(style="thin", color="D1D5DB")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    for cell in ws[linha_cabecalho]:
        cell.fill = preenchimento_cabecalho
        cell.font = fonte_cabecalho
        cell.border = borda
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for linha in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in linha:
            cell.border = borda
            cell.alignment = Alignment(vertical="center")

        tipo = linha[2].value

        if tipo == "Entrada":
            linha[2].fill = PatternFill("solid", fgColor="DCFCE7")
            linha[2].font = Font(color="166534", bold=True)

        if tipo == "Saída":
            linha[2].fill = PatternFill("solid", fgColor="FEE2E2")
            linha[2].font = Font(color="991B1B", bold=True)

    larguras = {
        "A": 24,
        "B": 32,
        "C": 16,
        "D": 26,
        "E": 18
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[4].height = 24

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{ws.max_row}"

    if ws.max_row >= 5:
        tabela = Table(displayName="TabelaRelatorioEstoque", ref=f"A4:E{ws.max_row}")

        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

    arquivo = "relatorio_estoque.xlsx"
    wb.save(arquivo)

    conexao.close()

    return send_file(arquivo, as_attachment=True)

@app.route("/limpar_relatorio")
def limpar_relatorio():
    if not verificar_login():
        return redirect("/login")

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("DELETE FROM movimentacoes")

    conexao.commit()
    conexao.close()

    return redirect("/relatorio")

@app.route("/excluir/<int:id>")
def excluir(id):
    if not verificar_login():
        return redirect("/login")

    conexao = conectar_banco()
    cursor = conexao.cursor()

    cursor.execute("DELETE FROM produtos WHERE id = ?", (id,))

    conexao.commit()
    conexao.close()

    return redirect("/")

app.run(debug=True)